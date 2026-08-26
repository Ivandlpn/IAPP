#!/usr/bin/env python3
"""
Lo aprovechable de los tres documentos de red de datos de seguridad de TOVAL.

Son documentos de arquitectura de red: no bajan a nivel de cámara, así que lo
único que se puede volcar al listado son campos deducibles del propio diseño
(la tecnología, que va por codificadores) y campos que ya están en el fichero
pero mal colocados. El resto es verificación.
"""
import json, os, re, sys, ipaddress, collections
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

INICIO, FIN = 2501, 4467
VERDE = PatternFill("solid", fgColor="FFD9EAD3")
GRIS = PatternFill("solid", fgColor="FFEFEFEF")
AZUL = PatternFill("solid", fgColor="FFDCE6F1")
FINO = Border(*[Side(style="thin")] * 4)
letra = lambda i: chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)


def hoja(wb, titulo, cab, filas, anchos, primera=None, congelar="A2"):
    ws = wb.create_sheet(titulo)
    for j, h in enumerate(cab, 1):
        c = ws.cell(1, j, h)
        c.font = Font(bold=True)
        c.fill = GRIS if (primera and j <= primera[0]) else (VERDE if (primera and j <= primera[1]) else AZUL)
        c.border = FINO
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[letra(j)].width = anchos[j - 1]
    ws.row_dimensions[1].height = 42
    for i, fila in enumerate(filas, 2):
        for j, v in enumerate(fila, 1):
            c = ws.cell(i, j, v)
            c.border = FINO
            c.alignment = Alignment(horizontal="left" if j <= 2 else "center")
    ws.freeze_panes = congelar
    ws.auto_filter.ref = f"A1:{letra(len(cab))}{len(filas) + 1}"
    return ws


def main():
    datos, v31, salida = sys.argv[1], sys.argv[2], sys.argv[3]
    zonas = json.load(open(os.path.join(datos, "tov_zonas.json")))
    redes = [(ipaddress.ip_network(k), v["zona"], v["vlan"]) for k, v in zonas.items()]

    ws31 = openpyxl.load_workbook(v31, data_only=True).active
    pegar, verif = [], []
    for r in range(INICIO, FIN + 1):
        v = [ws31.cell(r, c).value for c in range(1, 32)]
        pegar.append([r, v[1], "ANALÓGICA", v[1], "SEGURIDAD"])
        zona = vlan = None
        try:
            ip = ipaddress.ip_address(str(v[20]))
            for n, z, vl in redes:
                if ip in n:
                    zona, vlan = z, vl
                    break
        except ValueError:
            pass
        num = re.match(r"^(\d{1,2})\s*[–-]", zona or "")
        if not zona:
            estado = "IP fuera de las redes del plan"
        elif not num:
            estado = "zona propia (estación o base de mantenimiento)"
        elif str(v[27]) == num.group(1):
            estado = "coincide"
        else:
            estado = "el anillo del listado no es el de la red IP"
        verif.append([r, v[1], str(v[20]), v[27], zona, vlan, estado])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QUÉ SE PUEDE PEGAR"
    lineas = [
        ("Los tres documentos de red de TOVAL no bajan a nivel de cámara", True),
        ("", False),
        ("Son el plan de direccionamiento, la arquitectura y la descripción general de la "
         "red de datos de seguridad. Definen anillos, subredes, VLANs y routers, pero "
         "ninguna tabla asigna equipo por equipo.", False),
        ("", False),
        ("Lo que sí se puede rellenar, en la hoja PEGAR — filas 2501 a 4467:", True),
        ("   C2:C1968  →  D2501    TECNOLOGÍA = ANALÓGICA", False),
        ("   D2:D1968  →  H2501    NOMBRE DISPOSITIVO", False),
        ("   E2:E1968  →  S2501    FUNCIÓN CÁMARA = SEGURIDAD", False),
        ("", False),
        ("Por qué ANALÓGICA: los documentos hablan siempre de «Videovigilancia (Códecs)» y "
         "de «Codificadores/Decodificadores del Sistema de Vídeo», los modelos del listado "
         "son analógicos (DINION 2XF PAL 540LTV, VG4-513ECS, FLEXIDOMO VDN) y 829 IPs del "
         "tramo están compartidas por exactamente dos cámaras, que es el patrón de dos "
         "canales por codificador.", False),
        ("", False),
        ("NOMBRE DISPOSITIVO es una copia de REFERENCIA, como en el resto del fichero. "
         "No necesita los PDF.", False),
        ("", False),
        ("La hoja VERIFICACIÓN contrasta el anillo de cada cámara con la subred a la que "
         "pertenece su IP según el plan. No es para pegar.", False),
    ]
    for i, (t, b) in enumerate(lineas, 1):
        c = ws.cell(i, 1, t)
        c.font = Font(bold=b, size=12 if b else 11)
        c.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 112

    hoja(wb, "PEGAR",
         ["Fila v3.1", "REFERENCIA (comprobación)",
          f"C → pegar en D{INICIO}\nTECNOLOGÍA",
          f"D → pegar en H{INICIO}\nNOMBRE DISPOSITIVO",
          f"E → pegar en S{INICIO}\nFUNCIÓN CÁMARA"],
         pegar, [9, 26, 18, 26, 18], primera=(2, 5), congelar="C2")

    hoja(wb, "VERIFICACIÓN",
         ["Fila v3.1", "REFERENCIA", "DIRECCIÓN IP", "ANILLO en el listado",
          "Zona según el plan de direccionamiento", "VLAN", "Resultado"],
         verif, [9, 26, 16, 16, 40, 8, 42], primera=(2, 2))

    resumen = collections.Counter(x[6] for x in verif)
    hoja(wb, "REDES DEL PLAN",
         ["Red", "Zona", "VLAN"],
         sorted([[k, v["zona"], v["vlan"]] for k, v in zonas.items()],
                key=lambda x: ipaddress.ip_network(x[0]).network_address),
         [20, 44, 8], primera=(2, 3))

    wb.save(salida)
    print(f"{len(pegar)} filas | verificación: {dict(resumen)}")


if __name__ == "__main__":
    main()
