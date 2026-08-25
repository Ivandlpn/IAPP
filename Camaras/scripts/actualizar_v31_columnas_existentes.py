#!/usr/bin/env python3
"""
Vuelca la información de los dos documentos de proyecto (ALBALI y TOVAL) sobre
el listado unificado v3.1 SIN crear columnas nuevas: todo va a alguna de las 31
columnas que ya existen, o se queda fuera y se documenta como tal.

Salida: Actualizacion_v3.1_columnas_existentes.xlsx
"""
import json, os, re, sys, collections
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

ALBALI = (1989, 2500)
TOVAL = (2501, 4467)
MASCARA = {"24": 255255255000, "25": 255255255128, "26": 255255255192,
           "27": 255255255224, "28": 255255255240}

AMARILLO = PatternFill("solid", fgColor="FFFFF200")
VERDE = PatternFill("solid", fgColor="FFD9EAD3")
GRIS = PatternFill("solid", fgColor="FFEFEFEF")
ROJO = PatternFill("solid", fgColor="FFF4CCCC")
FINO = Border(*[Side(style="thin")] * 4)

clave = lambda s: re.sub(r"[^A-Z0-9]", "", str(s).upper())
sitio = lambda s: "_".join(str(s).split("_")[:3])
es_na = lambda v: str(v).strip() == "#N/A"
letra = lambda i: chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)


def hoja(wb, titulo, cabeceras, filas, anchos, relleno=VERDE, congelar="A2"):
    ws = wb.create_sheet(titulo)
    for j, h in enumerate(cabeceras, 1):
        c = ws.cell(1, j, h)
        c.font, c.fill, c.border = Font(bold=True), relleno, FINO
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[letra(j)].width = anchos[j - 1]
    for i, fila in enumerate(filas, 2):
        for j, v in enumerate(fila, 1):
            c = ws.cell(i, j, v)
            c.border = FINO
            c.alignment = Alignment(horizontal="left" if j <= 2 else "center",
                                    vertical="top", wrap_text=j > 2 and anchos[j - 1] > 30)
    ws.freeze_panes = congelar
    ws.auto_filter.ref = f"A1:{letra(len(cabeceras))}{len(filas) + 1}"
    return ws


def texto(ws, lineas, ancho=112):
    for i, (t, negrita) in enumerate(lineas, 1):
        c = ws.cell(i, 1, t)
        c.font = Font(bold=negrita, size=12 if negrita else 11)
        c.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = ancho


def main():
    datos, v31, salida = sys.argv[1], sys.argv[2], sys.argv[3]
    cam = json.load(open(os.path.join(datos, "pdf_camaras.json")))
    gw_anillo = json.load(open(os.path.join(datos, "gw_anillo.json")))
    grab = json.load(open(os.path.join(datos, "toval_grabadores.json")))
    sit2gr = json.load(open(os.path.join(datos, "toval_asignacion.json")))["sit2gr"]

    por_nombre, por_ip = {}, {}
    for c in cam:
        if c.get("etq"):
            por_nombre.setdefault(clave(c["etq"]), c)
            por_ip.setdefault(c["ip"], c)

    ws31 = openpyxl.load_workbook(v31, data_only=True).active
    fila = lambda r: [ws31.cell(r, c).value for c in range(1, 32)]

    # ------------------------------------------------------------- ALBALI
    # cada celda lleva el estado final: el dato del proyecto si lo hay, y si no
    # el valor que ya tiene el v3.1 (vacío cuando ese valor es #N/A). Así pegar
    # el bloque entero nunca borra un dato bueno.
    def final(nuevo, actual):
        if nuevo not in (None, ""):
            return nuevo
        return None if (actual in (None, "") or es_na(actual)) else actual

    alb, fuera_alb = [], 0
    for r in range(ALBALI[0], ALBALI[1] + 1):
        v = fila(r)
        c = por_nombre.get(clave(v[1])) or por_ip.get(str(v[20]))
        if not c:
            fuera_alb += 1
            c = {}
        alb.append([r, v[1],
                    final(c.get("emp"), v[12]),
                    final(c.get("pk"), v[13]),
                    final(MASCARA.get(c["masc"], "/" + c["masc"]) if c else None, v[21]),
                    final(c.get("gw"), v[22]),
                    final(gw_anillo.get(c["gw"]) if c else None, v[27]),
                    final(c.get("g1"), v[28]),
                    final(c.get("ipg1"), v[28])])

    # -------------------------------------------------------------- TOVAL
    tov, unico, ambiguo, sin_sitio = [], 0, 0, 0
    for r in range(TOVAL[0], TOVAL[1] + 1):
        v = fila(r)
        gs = sit2gr.get(sitio(v[1])) or sit2gr.get(sitio(v[23])) or []
        g = gs[0] if len(gs) == 1 else None
        if g:
            unico += 1
        elif gs:
            ambiguo += 1
        else:
            sin_sitio += 1
        tov.append([r, v[1],
                    final(g, v[28]),
                    final(grab.get(g, {}).get("ip") if g else None, v[28]),
                    " / ".join(gs) if len(gs) > 1 else None])

    # ------------------------------------------------------------- libro
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CÓMO PEGARLO"
    texto(ws, [
        ("Actualización del v3.1 sin columnas nuevas", True),
        ("", False),
        ("Todo lo que sigue va a columnas que ya existen. Lo que no cabe en ninguna "
         "se queda fuera y está listado en la hoja DATOS QUE FALTAN.", False),
        ("Pega siempre con Pegado especial > Valores.", False),
        ("", False),
        ("BLOQUE ALBALI — hoja «PEGAR ALBALI», filas 1989 a 2500 del v3.1", True),
        ("   C2:C513   →  M1989    UBICACIÓN", False),
        ("   D2:D513   →  N1989    NUMERACION SOBRE PLANO  (el PK absoluto del proyecto)", False),
        ("   E2:F513   →  V1989    MÁSCARA DE RED + PUERTA DE ENLACE", False),
        ("   G2:G513   →  AB1989   ANILLO", False),
        ("   H2:H513   →  AC1989   SERVIDOR DE GRABACIÓN  (nombre del grabador)", False),
        ("   Si prefieres dejar la IP que ya puso tu jefe en AC, usa I2:I513 en su lugar "
         "y no pegues H.", False),
        ("", False),
        ("BLOQUE TOVAL — hoja «PEGAR TOVAL», filas 2501 a 4467 del v3.1", True),
        ("   C2:C1968  →  AC2501   SERVIDOR DE GRABACIÓN  (nombre del grabador)", False),
        ("   o bien D2:D1968 si prefieres la IP del grabador. Nunca las dos.", False),
        ("   La columna E es informativa: los grabadores candidatos cuando el "
         "emplazamiento tiene más de uno.", False),
        ("", False),
        ("Antes de pegar, comprueba que la columna B de cada hoja coincide con la "
         "columna B del v3.1 en esas filas.", False),
    ])

    hoja(wb, "MAPEO",
         ["Documento", "Dato del documento", "Columna del v3.1", "Col.", "Filas que rellena", "Nota"],
         [
             ["ALBALI", "EMPLAZAMIENTO", "UBICACIÓN", "M", 49 + 6,
              "49 estaban en #N/A y 6 estaban truncadas"],
             ["ALBALI", "PK ABSOLUTO", "NUMERACION SOBRE PLANO", "N", 468,
              "La columna está vacía en ALBALI y es la referencia de posición de la cámara"],
             ["ALBALI", "MÁSCARA (/27, /26)", "MÁSCARA DE RED", "V", 49,
              "Convertida al formato del fichero: 255255255224 / 255255255192"],
             ["ALBALI", "GATEWAY", "PUERTA DE ENLACE (GATEWEY)", "W", 49, "Estaban en #N/A"],
             ["ALBALI", "Anillo (deducido del gateway)", "ANILLO", "AB", 3, "Estaban en #N/A"],
             ["ALBALI", "GRABADOR PPAL", "SERVIDOR DE GRABACIÓN", "AC", 468,
              "Nombre del grabador; alternativa: la IP, que es lo que hay hoy"],
             ["ALBALI", "IP GRABADOR SEC", "— ninguna —", "", 0, "NO CABE: no hay columna de grabador secundario"],
             ["ALBALI", "GRABADOR SEC", "— ninguna —", "", 0, "NO CABE: ídem"],
             ["ALBALI", "STR1 resolución / ips / compresión", "— ninguna —", "", 0,
              "NO CABE: no hay columnas de parámetros de vídeo"],
             ["ALBALI", "STR2 resolución / ips / compresión", "— ninguna —", "", 0, "NO CABE: ídem"],
             ["TOVAL", "GRABADOR del emplazamiento", "SERVIDOR DE GRABACIÓN", "AC", unico,
              "Solo donde el emplazamiento tiene un único grabador"],
             ["TOVAL", "IP del grabador", "SERVIDOR DE GRABACIÓN", "AC", unico,
              "Alternativa al nombre, no las dos"],
             ["TOVAL", "Servidor VRM (1, 2 o 3)", "— ninguna —", "", 0,
              "NO CABE: no hay columna de servidor de gestión"],
             ["TOVAL", "Asignación codificador → grabador", "CODIFICADOR/PUERTO", "E", 0,
              "NO SE PUEDE APLICAR: la columna E está vacía en TOVAL. Es el dato que falta"],
         ],
         [12, 38, 30, 8, 16, 58])

    hoja(wb, "PEGAR ALBALI",
         ["Fila v3.1", "REFERENCIA (comprobación)", "UBICACIÓN → M1989",
          "NUMERACION SOBRE PLANO → N1989", "MÁSCARA DE RED → V1989",
          "PUERTA DE ENLACE → W1989", "ANILLO → AB1989",
          "SERVIDOR DE GRABACIÓN → AC1989 (nombre)",
          "alternativa: IP del grabador"],
         alb, [10, 26, 40, 16, 18, 18, 10, 26, 20], GRIS, "C2")

    hoja(wb, "PEGAR TOVAL",
         ["Fila v3.1", "REFERENCIA (comprobación)",
          "SERVIDOR DE GRABACIÓN → AC2501 (nombre)", "alternativa: IP del grabador",
          "Grabadores candidatos (informativo, NO pegar)"],
         tov, [10, 26, 30, 20, 44], GRIS, "C2")

    hoja(wb, "DATOS QUE FALTAN",
         ["Prioridad", "Qué falta", "A quién / dónde pedirlo", "Qué desbloquea"],
         [
             [1, "Columna CODIFICADOR/PUERTO de TOVAL: a qué codificador y puerto "
                 "va cada una de las 1.967 cámaras",
              "Configuración de los codificadores Bosch, o el as-built de TOVAL",
              f"Cerrar SERVIDOR DE GRABACIÓN de las {ambiguo + sin_sitio} cámaras de TOVAL que "
              "hoy quedan sin grabador, con principal y fail-over exactos"],
             [2, "SWITCH y BOCA de las 512 cámaras de ALBALI",
              "Esquemas de red de ALBALI; el documento de CCTV no los trae",
              "Columnas X y AA, hoy vacías en todo el tramo"],
             [3, "Datos de las 44 cámaras de ALBALI que no están en el proyecto "
                 "(6AA_D_46xx, rango 24.17.145.x)",
              "Revisión posterior a la Rev. A de 07/2013 del proyecto ALBALI",
              "UBICACIÓN, MÁSCARA, GATEWAY, ANILLO y SERVIDOR de esas 44 filas"],
             [4, "Números de serie reales del bloque SOAV",
              "El fichero de origen, antes de que Excel los convirtiera a número",
              "639 series que hoy se han quedado en solo 79 valores distintos"],
             [5, "Ámbito de las 1.623 filas de SOAV sin etiqueta de tramo",
              "Quien montó el listado original",
              "Columnas A y AD, para poder segmentar el fichero entero por tramo"],
             [6, "DESCRIPCIÓN y FUNCIÓN CÁMARA de ALBALI",
              "Inventario de ALBALI",
              "Columnas C y S, vacías en las 512 filas"],
             [7, "Las 77 cámaras de la Estación de Alicante que están en el proyecto "
                 "ALBALI y no en el listado",
              "Confirmar si se instalaron y no se exportaron, o si no existen",
              "Completar el inventario del tramo"],
         ],
         [10, 52, 46, 60])

    wb.save(salida)
    print(f"ALBALI: {512 - fuera_alb} con datos, {fuera_alb} sin PDF | "
          f"TOVAL: {unico} con grabador único, {ambiguo} ambiguas, {sin_sitio} sin sitio")


if __name__ == "__main__":
    main()
