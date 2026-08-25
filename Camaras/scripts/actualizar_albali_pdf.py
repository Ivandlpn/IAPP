#!/usr/bin/env python3
"""
Cruza el bloque ALBALI del listado unificado v3.1 con el proyecto
ALBALI-AV-VCA-M700-SAL-002 (subsistema CCTV) y genera:

  1. ALBALI_cambios_desde_PDF.xlsx  - qué celda cambiar y por qué valor
  2. Camaras_unificado_v3.1_ALBALI_actualizado.xlsx - el bloque ya corregido

El cruce se hace por REFERENCIA (nombre del equipo) y, cuando el nombre no
coincide, por DIRECCIÓN IP.
"""
import json, re, sys, os
from copy import copy
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

MASCARA = {"24": 255255255000, "25": 255255255128, "26": 255255255192,
           "27": 255255255224, "28": 255255255240}

# columna destino (1-based) para cada campo del PDF
DESTINO = {
    "emp":  (13, "UBICACIÓN"),
    "masc": (22, "MÁSCARA DE RED"),
    "gw":   (23, "PUERTA DE ENLACE (GATEWEY)"),
    "ipg1": (29, "SERVIDOR DE GRABACIÓN"),
    "anillo": (28, "ANILLO"),
}   # el PK no se vuelca aquí: va en su propia columna nueva (ver EXTRA)
# campos del PDF que no tienen columna en el formato actual
EXTRA = [("pk", "PK ABSOLUTO"), ("g1", "GRABADOR PPAL"), ("ipg2", "IP GRABADOR SEC"),
         ("g2", "GRABADOR SEC"), ("s1r", "STR1 RESOLUCIÓN"), ("s1i", "STR1 IPS"),
         ("s1c", "STR1 COMPRESIÓN"), ("s2r", "STR2 RESOLUCIÓN"), ("s2i", "STR2 IPS"),
         ("s2c", "STR2 COMPRESIÓN")]

clave = lambda s: re.sub(r"[^A-Z0-9]", "", str(s).upper())
es_na = lambda v: str(v).strip() == "#N/A"
vacio = lambda v: v in (None, "") or es_na(v)


def letra(i):
    return chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)


def cargar_pdf(dir_datos):
    cam = json.load(open(os.path.join(dir_datos, "pdf_camaras.json")))
    gw_anillo = json.load(open(os.path.join(dir_datos, "gw_anillo.json")))
    for c in cam:
        c["anillo"] = gw_anillo.get(c["gw"])
        c["masc_excel"] = MASCARA.get(c["masc"], "/" + c["masc"])
    por_nombre, por_ip = {}, {}
    for c in cam:
        if c.get("etq"):
            por_nombre.setdefault(clave(c["etq"]), c)
            por_ip.setdefault(c["ip"], c)
    return cam, por_nombre, por_ip


def emparejar(filas, por_nombre, por_ip):
    """Devuelve (fila, valores, ficha_pdf, via) por cada cámara del listado."""
    for r, v in filas:
        c = por_nombre.get(clave(v[1]))
        if c:
            yield r, v, c, "nombre"
            continue
        c = por_ip.get(str(v[20]))
        yield r, v, c, ("IP" if c else None)


def calcular_cambios(pares):
    cambios = []
    for r, v, c, via in pares:
        if not c:
            continue
        for campo, (idx, nombre) in DESTINO.items():
            nuevo = c["masc_excel"] if campo == "masc" else c.get(campo)
            if nuevo in (None, ""):
                continue
            actual = v[idx - 1]
            if vacio(actual):
                motivo = "Rellenar #N/A" if es_na(actual) else "Rellenar vacío"
            elif str(actual).strip() != str(nuevo).strip():
                motivo = "Corregir"
            else:
                continue
            cambios.append({
                "fila": r, "col": letra(idx), "columna": nombre,
                "referencia": v[1], "actual": "" if actual is None else actual,
                "nuevo": nuevo, "motivo": motivo, "cruce": via, "pag": c["pag"],
            })
    return cambios


# ---------------------------------------------------------------- escritura
CAB = PatternFill("solid", fgColor="FFFFF200")
BORDE = Border(*[Side(style="thin")] * 4)


def hoja(wb, titulo, cabeceras, filas, anchos=None):
    ws = wb.create_sheet(titulo) if wb.sheetnames != ["Sheet"] else wb.active
    ws.title = titulo
    for j, h in enumerate(cabeceras, 1):
        c = ws.cell(1, j, h)
        c.font, c.fill, c.border = Font(bold=True), CAB, BORDE
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, fila in enumerate(filas, 2):
        for j, val in enumerate(fila, 1):
            c = ws.cell(i, j, val)
            c.border = BORDE
            c.alignment = Alignment(horizontal="left" if j <= 4 else "center")
    for j, w in enumerate(anchos or [], 1):
        ws.column_dimensions[letra(j)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{letra(len(cabeceras))}{len(filas) + 1}"
    return ws


def main():
    datos = sys.argv[1]      # carpeta con pdf_camaras.json y gw_anillo.json
    v31 = sys.argv[2]        # Cámaras_unificado_v3.1_SOAV.xlsx
    salida = sys.argv[3]

    cam, por_nombre, por_ip = cargar_pdf(datos)
    wb31 = openpyxl.load_workbook(v31)
    ws31 = wb31.active
    filas = [(r, [ws31.cell(r, c).value for c in range(1, 32)]) for r in range(1989, 2501)]
    pares = list(emparejar(filas, por_nombre, por_ip))
    cambios = calcular_cambios(pares)

    usados = {p[2]["etq"] for p in pares if p[2]}
    sin_pdf = [(r, v[1], v[20]) for r, v, c, _ in pares if not c]
    sobran = [c for c in cam if c.get("etq") and c["etq"] not in usados]
    reserva = [c for c in cam if not c.get("etq")]

    # ---- libro de cambios
    wb = openpyxl.Workbook()
    hoja(wb, "Cambios",
         ["Fila", "Col.", "Columna", "REFERENCIA", "Valor actual", "Valor nuevo",
          "Motivo", "Cruzado por", "Pág. PDF"],
         [[c["fila"], c["col"], c["columna"], c["referencia"], c["actual"],
           c["nuevo"], c["motivo"], c["cruce"], c["pag"]] for c in cambios],
         [8, 7, 28, 24, 42, 42, 15, 12, 10])

    hoja(wb, "Datos nuevos",
         ["Fila", "REFERENCIA"] + [n for _, n in EXTRA],
         [[r, v[1]] + [c.get(k) for k, _ in EXTRA] for r, v, c, _ in pares if c],
         [8, 24] + [16] * len(EXTRA))

    distintos = [(r, v[1], c["etq"], v[20], c["emp"]) for r, v, c, via in pares
                 if c and via == "IP"]
    hoja(wb, "Nombres que no coinciden",
         ["Fila", "REFERENCIA en el listado", "Etiqueta en el PDF", "DIRECCIÓN IP",
          "Emplazamiento (PDF)"], distintos, [8, 26, 26, 18, 40])

    hoja(wb, "Sin dato en el PDF",
         ["Fila", "REFERENCIA", "DIRECCIÓN IP"], sin_pdf, [8, 26, 18])

    hoja(wb, "En el PDF y no en el listado",
         ["Etiqueta", "Emplazamiento", "Descripción", "IP", "Máscara", "Gateway",
          "PK", "Grabador ppal", "Pág. PDF"],
         [[c["etq"], c["emp"], c["desc"], c["ip"], "/" + c["masc"], c["gw"],
           c["pk"], c["g1"], c["pag"]] for c in sobran],
         [24, 34, 40, 16, 10, 16, 12, 20, 10])

    hoja(wb, "IPs de reserva",
         ["Emplazamiento", "Descripción", "IP", "Máscara", "Gateway", "PK", "Pág. PDF"],
         [[c["emp"], c["desc"], c["ip"], "/" + c["masc"], c["gw"], c["pk"], c["pag"]]
          for c in reserva],
         [34, 20, 16, 10, 16, 12, 10])
    wb.save(os.path.join(salida, "ALBALI_cambios_desde_PDF.xlsx"))

    # ---- ALBALI actualizado, mismo formato + columnas nuevas al final
    out = openpyxl.Workbook()
    ws = out.active
    ws.title = "Listado cámaras LAV"
    for c in range(1, 32):
        src, dst = ws31.cell(1, c), ws.cell(1, c)
        dst.value = src.value
        for a in ("font", "fill", "border", "alignment"):
            setattr(dst, a, copy(getattr(src, a)))
    ws.row_dimensions[1].height = ws31.row_dimensions[1].height
    for k, dim in ws31.column_dimensions.items():
        if dim.width:
            ws.column_dimensions[k].width = dim.width
    for j, (_, nombre) in enumerate(EXTRA, 32):
        c = ws.cell(1, j, nombre)
        c.font, c.fill, c.border = Font(bold=True), CAB, BORDE
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[letra(j)].width = 18

    porfila = {}
    for c in cambios:
        porfila.setdefault(c["fila"], {})[c["col"]] = c["nuevo"]
    fuente, centro = Font(name="Calibri", size=11), Alignment(horizontal="center")
    punto = Border(*[Side(style="dotted")] * 4)
    for i, (r, v, ficha, _) in enumerate(pares, 2):
        nuevos = porfila.get(r, {})
        for j in range(1, 32):
            cel = ws.cell(i, j)
            val = nuevos.get(letra(j), v[j - 1])
            cel.value = None if es_na(val) else val
            cel.font, cel.alignment, cel.border = fuente, centro, punto
            if j == 17 and cel.value is not None:
                cel.number_format = "@"
        for j, (k, _) in enumerate(EXTRA, 32):
            cel = ws.cell(i, j, ficha.get(k) if ficha else None)
            cel.font, cel.alignment, cel.border = fuente, centro, punto
    ws.auto_filter.ref = f"A1:{letra(31 + len(EXTRA))}{len(pares) + 1}"
    ws.freeze_panes = "A2"
    out.save(os.path.join(salida, "Camaras_unificado_v3.1_ALBALI_actualizado.xlsx"))

    print(f"cambios: {len(cambios)} | sin dato en PDF: {len(sin_pdf)} | "
          f"en PDF y no en el listado: {len(sobran)} | IPs de reserva: {len(reserva)}")


if __name__ == "__main__":
    main()
