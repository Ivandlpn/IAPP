#!/usr/bin/env python3
"""
Un fichero por tramo, con una sola hoja de datos, alineada fila a fila con el
listado unificado v3.1. Cada columna lleva escrito en la cabecera la celda del
v3.1 donde se pega, y contiene el estado final de esa columna: el dato del
proyecto donde lo hay y el valor actual donde no, para que pegar el bloque
entero nunca borre nada bueno.
"""
import json, os, re, sys
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

MASCARA = {"24": 255255255000, "25": 255255255128, "26": 255255255192,
           "27": 255255255224, "28": 255255255240}
VERDE = PatternFill("solid", fgColor="FFD9EAD3")
GRIS = PatternFill("solid", fgColor="FFEFEFEF")
AZUL = PatternFill("solid", fgColor="FFDCE6F1")
FINO = Border(*[Side(style="thin")] * 4)

clave = lambda s: re.sub(r"[^A-Z0-9]", "", str(s).upper())
sitio = lambda s: "_".join(str(s).split("_")[:3])
es_na = lambda v: str(v).strip() == "#N/A"
letra = lambda i: chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)


def final(nuevo, actual):
    if nuevo not in (None, ""):
        return nuevo
    return None if (actual in (None, "") or es_na(actual)) else actual


def escribir(ruta, titulo, cabeceras, filas, anchos, consulta_desde):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo
    for j, h in enumerate(cabeceras, 1):
        c = ws.cell(1, j, h)
        c.font = Font(bold=True)
        c.fill = GRIS if j <= 2 else (AZUL if j >= consulta_desde else VERDE)
        c.border = FINO
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[letra(j)].width = anchos[j - 1]
    ws.row_dimensions[1].height = 46
    for i, fila in enumerate(filas, 2):
        for j, v in enumerate(fila, 1):
            c = ws.cell(i, j, v)
            c.border = FINO
            c.alignment = Alignment(horizontal="left" if j <= 2 else "center")
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{letra(len(cabeceras))}{len(filas) + 1}"
    wb.save(ruta)


def main():
    datos, v31, destino = sys.argv[1], sys.argv[2], sys.argv[3]
    cam = json.load(open(os.path.join(datos, "pdf_camaras.json")))
    gw_anillo = json.load(open(os.path.join(datos, "gw_anillo.json")))
    grab = json.load(open(os.path.join(datos, "toval_grabadores.json")))
    sit2gr = json.load(open(os.path.join(datos, "toval_asignacion.json")))["sit2gr"]

    por_nombre, por_ip = {}, {}
    for c in cam:
        if c.get("etq"):
            por_nombre.setdefault(clave(c["etq"]), c)
            por_ip.setdefault(c["ip"], c)

    ws31 = openpyxl.load_workbook(v31, data_only=True).active
    fila = lambda r: [ws31.cell(r, c).value for c in range(1, 32)]

    # -------------------------------------------------------------- ALBALI
    filas = []
    for r in range(1989, 2501):
        v = fila(r)
        c = por_nombre.get(clave(v[1])) or por_ip.get(str(v[20])) or {}
        filas.append([
            r, v[1],
            final(c.get("emp"), v[12]),                                    # M
            final(c.get("pk"), v[13]),                                     # N
            final(MASCARA.get(c["masc"]) if c else None, v[21]),           # V
            final(c.get("gw"), v[22]),                                     # W
            final(gw_anillo.get(c["gw"]) if c else None, v[27]),           # AB
            final(c.get("g1"), v[28]),                                     # AC
            c.get("ipg1"), c.get("g2"), c.get("ipg2"),
            c.get("s1r"), c.get("s1i"), c.get("s2r"), c.get("s2i"),
        ])
    escribir(os.path.join(destino, "ALBALI_pegar_en_v3.1.xlsx"), "PEGAR EN EL v3.1",
             ["Fila v3.1", "REFERENCIA\n(comprobación)",
              "C → pegar en M1989\nUBICACIÓN",
              "D → pegar en N1989\nNUMERACION SOBRE PLANO",
              "E → pegar en V1989\nMÁSCARA DE RED",
              "F → pegar en W1989\nPUERTA DE ENLACE",
              "G → pegar en AB1989\nANILLO",
              "H → pegar en AC1989\nSERVIDOR DE GRABACIÓN",
              "solo consulta\nIP grabador ppal", "solo consulta\nGrabador secundario",
              "solo consulta\nIP grabador sec", "solo consulta\nSTR1 resol.",
              "solo consulta\nSTR1 ips", "solo consulta\nSTR2 resol.",
              "solo consulta\nSTR2 ips"],
             filas, [9, 22, 38, 16, 16, 16, 10, 22, 16, 20, 16, 12, 10, 12, 10], 9)

    # --------------------------------------------------------------- TOVAL
    filas = []
    for r in range(2501, 4468):
        v = fila(r)
        gs = sit2gr.get(sitio(v[1])) or sit2gr.get(sitio(v[23])) or []
        g = gs[0] if len(gs) == 1 else None
        d = grab.get(g, {}) if g else {}
        filas.append([r, v[1], final(g, v[28]), d.get("ip"), d.get("desc"),
                      int(d["vrm"]) if d.get("vrm") else None,
                      " / ".join(gs) if len(gs) > 1 else None])
    escribir(os.path.join(destino, "TOVAL_pegar_en_v3.1.xlsx"), "PEGAR EN EL v3.1",
             ["Fila v3.1", "REFERENCIA\n(comprobación)",
              "C → pegar en AC2501\nSERVIDOR DE GRABACIÓN",
              "solo consulta\nIP del grabador", "solo consulta\nEmplazamiento grabador",
              "solo consulta\nVRM", "solo consulta\nCandidatos si hay varios"],
             filas, [9, 22, 24, 16, 30, 8, 40], 4)

    print("ALBALI_pegar_en_v3.1.xlsx y TOVAL_pegar_en_v3.1.xlsx generados")


if __name__ == "__main__":
    main()
