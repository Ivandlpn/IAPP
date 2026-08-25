#!/usr/bin/env python3
"""
Genera tres listados de camaras (ALBALI, CHATO, TOVAL) con el formato de
'Camaras_unificado_v2_SOAV.xlsx': misma cabecera de 31 columnas, mismos anchos,
mismo estilo, autofiltro y panel inmovilizado.

La columna AD (30) identifica el ambito: ALBALI / CHATO / TOVAL.

Uso:  python3 generar_listados.py <dir_entrada> <dir_salida>
"""
import sys, os, re
from copy import copy
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

PLANTILLA = "5e8a8b75-C_maras_unificado_v2_SOAV.xlsx"
F_ALBALI  = "8192cdad-Listado_C_maras_ALBALI.xlsx"
F_CHATO   = "4a87e5cc-Listado_C_maras_CHATO.xlsx"
F_TOVAL   = "b92479fd-Listado_C_maras_TOVAL.xlsx"

NCOL = 31
COL = {  # nombre logico -> indice 1-based en el formato destino
    "AMBITO": 1, "REFERENCIA": 2, "DESCRIPCION": 3, "TECNOLOGIA": 4,
    "CODIFICADOR": 5, "PUERTO": 6, "COD_HARDWARE": 7, "NOMBRE_DISPOSITIVO": 8,
    "GRAB_MILESTONE": 9, "NOMBRE_MILESTONE": 10, "GRAB_PUERTO": 11, "PLANTA": 12,
    "UBICACION": 13, "NUM_PLANO": 14, "TIPO_CAMARA": 15, "MODELO": 16,
    "SERIE": 17, "MAC": 18, "FUNCION": 19, "FIRMWARE": 20, "IP": 21,
    "MASCARA": 22, "GATEWAY": 23, "SWITCH": 24, "SW_ALTO": 25, "SW_BAJO": 26,
    "BOCA": 27, "ANILLO": 28, "SERVIDOR": 29, "LAV": 30, "LIBRE": 31,
}


def norm(s):
    """Normaliza un titulo de columna para poder emparejarlo entre ficheros."""
    import unicodedata
    s = str(s or "").strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    return re.sub(r"[^A-Z0-9]", "", s)


def limpiar(v):
    """Quita saltos de linea y espacios sobrantes de los textos."""
    if isinstance(v, str):
        v = re.sub(r"\s*\n\s*", "", v).strip()
        return v or None
    return v


# --------------------------------------------------------------------------
# Lectura de las fuentes -> lista de dicts con las claves de COL
# --------------------------------------------------------------------------

def leer_albali(path):
    """Export 'SystemReport' de IndigoVision: secciones por modelo, cada una con
    su cabecera 'Friendly Name / IP Address / MAC Address / Version / Serial # / UDN'."""
    ws = openpyxl.load_workbook(path, data_only=True).active
    filas, modelo, tipo = [], None, None
    for r in range(1, ws.max_row + 1):
        a, b = ws.cell(r, 1).value, ws.cell(r, 2).value
        if a and b is None:                      # cabecera de seccion: "MODELO TIPO ..."
            modelo, _, tipo = str(a).strip().partition(" ")
            tipo = tipo.strip()
            continue
        if a and str(a).startswith("Friendly Name"):
            continue
        if not a or not b:
            continue
        filas.append({
            "REFERENCIA": limpiar(a),
            "NOMBRE_DISPOSITIVO": limpiar(a),
            "TECNOLOGIA": "Digital",             # todas las secciones son "... IP"
            "TIPO_CAMARA": tipo,
            "MODELO": modelo,
            "IP": limpiar(b),
            "MAC": (limpiar(ws.cell(r, 3).value) or "").upper() or None,
            "FIRMWARE": limpiar(ws.cell(r, 4).value),
            "SERIE": limpiar(ws.cell(r, 5).value),
            "COD_HARDWARE": limpiar(ws.cell(r, 6).value),   # UDN del dispositivo
            "LAV": "ALBALI",
        })
    return filas


def leer_formato_estandar(path, hoja, fila_cabecera, etiqueta):
    """CHATO y TOVAL ya vienen con la cabecera del formato destino: se emparejan
    las columnas por nombre normalizado."""
    ws = openpyxl.load_workbook(path, data_only=True)[hoja]
    cab = {}
    for c in range(1, ws.max_column + 1):
        t = norm(ws.cell(fila_cabecera, c).value)
        if t:
            cab[t] = c
    destino = {
        "REFERENCIA": "REFERENCIA", "DESCRIPCION": "DESCRIPCION",
        "TECNOLOGIA": "TECNOLOGIA", "CODIFICADOR": "CODIFICADORPUERTO",
        "PUERTO": "PUERTO", "COD_HARDWARE": "CODHARDWARE",
        "NOMBRE_DISPOSITIVO": "NOMBREDISPOSITIVO",
        "GRAB_MILESTONE": "GRABADORMILESTONE", "NOMBRE_MILESTONE": "NOMBREMILESTONE",
        "GRAB_PUERTO": "GRABADORPUERTO", "PLANTA": "PLANTA",
        "UBICACION": "UBICACION", "NUM_PLANO": "NUMERACIONSOBREPLANO",
        "TIPO_CAMARA": "TIPOCAMARA", "MODELO": "MODELO",
        "SERIE": "NUMERODESERIE", "MAC": "DIRECCIONMAC",
        "FUNCION": "FUNCIONCAMARA", "FIRMWARE": "FIRMWARE",
        "IP": "DIRECCIONIP", "MASCARA": "MASCARADERED",
        "SWITCH": "SWITCH", "SW_ALTO": "NOMBRESWITCHESQUEMAALTONIVEL",
        "SW_BAJO": "NOMBRESWITCHESQUEMABAJONIVEL", "BOCA": "BOCA",
        "ANILLO": "ANILLO", "SERVIDOR": "SERVIDORDEGRABACION",
    }
    gw = cab.get("PUERTADEENLACEGATEWEY") or cab.get("PUERTADEENLACEGATEWAY")
    filas = []
    for r in range(fila_cabecera + 1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(v not in (None, "") for v in vals):
            continue
        f = {"LAV": etiqueta}
        for k, t in destino.items():
            if t in cab:
                f[k] = limpiar(ws.cell(r, cab[t]).value)
        if gw:
            f["GATEWAY"] = limpiar(ws.cell(r, gw).value)
        filas.append(f)
    return filas


# --------------------------------------------------------------------------
# Escritura con el formato de la plantilla
# --------------------------------------------------------------------------

def escribir(filas, plantilla, salida, titulo_hoja):
    orig = openpyxl.load_workbook(plantilla).active
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo_hoja

    for c in range(1, NCOL + 1):
        src, dst = orig.cell(1, c), ws.cell(1, c)
        dst.value = src.value
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
    ws.row_dimensions[1].height = orig.row_dimensions[1].height
    for letra, dim in orig.column_dimensions.items():
        if dim.width:
            ws.column_dimensions[letra].width = dim.width

    fuente = Font(name="Calibri", size=11)
    centro = Alignment(horizontal="center")
    punto = Side(style="dotted")
    borde = Border(left=punto, right=punto, top=punto, bottom=punto)

    for i, f in enumerate(filas, start=2):
        for nombre, c in COL.items():
            cel = ws.cell(i, c)
            v = f.get(nombre)
            cel.value = v
            cel.font, cel.alignment, cel.border = fuente, centro, borde
            if nombre == "SERIE" and v is not None:
                cel.number_format = "@"          # serie siempre como texto

    ws.auto_filter.ref = f"A1:AD{len(filas) + 1}"
    ws.freeze_panes = "A2"
    wb.save(salida)
    return len(filas)


def main():
    ent = sys.argv[1] if len(sys.argv) > 1 else "."
    sal = sys.argv[2] if len(sys.argv) > 2 else "."
    os.makedirs(sal, exist_ok=True)
    plantilla = os.path.join(ent, PLANTILLA)

    trabajos = [
        ("ALBALI", leer_albali(os.path.join(ent, F_ALBALI))),
        ("CHATO", leer_formato_estandar(os.path.join(ent, F_CHATO), "Cámaras", 3, "CHATO")),
        ("TOVAL", leer_formato_estandar(os.path.join(ent, F_TOVAL), "Hoja1", 3, "TOVAL")),
    ]
    for etiqueta, filas in trabajos:
        destino = os.path.join(sal, f"Camaras_unificado_v2_{etiqueta}.xlsx")
        n = escribir(filas, plantilla, destino, "Listado cámaras LAV")
        print(f"{etiqueta:<7} {n:>5} cámaras -> {destino}")


if __name__ == "__main__":
    main()
