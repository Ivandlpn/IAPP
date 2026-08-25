#!/usr/bin/env python3
"""
Genera 'ALBALI_para_pegar.xlsx': los datos de ALBALI ya corregidos, ordenados
igual que las filas 1989-2500 del listado unificado v3.1 y agrupados en cuatro
bloques de columnas contiguas, para pegarlos de una vez en el fichero original.
"""
import sys, os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# (columna destino donde empieza el bloque, [cabeceras del bloque])
BLOQUES = [
    ("M", ["UBICACIÓN"]),
    ("V", ["MÁSCARA DE RED", "PUERTA DE ENLACE (GATEWEY)"]),
    ("AB", ["ANILLO", "SERVIDOR DE GRABACIÓN"]),
    ("AF", ["PK ABSOLUTO", "GRABADOR PPAL", "IP GRABADOR SEC", "GRABADOR SEC",
            "STR1 RESOLUCIÓN", "STR1 IPS", "STR1 COMPRESIÓN",
            "STR2 RESOLUCIÓN", "STR2 IPS", "STR2 COMPRESIÓN"]),
]
FILA_INICIAL = 1989

AMARILLO = PatternFill("solid", fgColor="FFFFF200")
VERDE = PatternFill("solid", fgColor="FFD9EAD3")
GRIS = PatternFill("solid", fgColor="FFEFEFEF")
FINO = Border(*[Side(style="thin")] * 4)


def letra(i):
    return chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)


def main():
    origen, salida = sys.argv[1], sys.argv[2]
    wa = openpyxl.load_workbook(origen, data_only=True).active
    cab = {wa.cell(1, c).value: c for c in range(1, wa.max_column + 1)}
    filas = [[wa.cell(r, c).value for c in range(1, wa.max_column + 1)]
             for r in range(2, wa.max_row + 1)]

    wb = openpyxl.Workbook()

    # ---- instrucciones
    ws = wb.active
    ws.title = "CÓMO PEGARLO"
    pasos = [
        ["Paso", "Qué copiar de la hoja DATOS", "Dónde pegarlo en el v3.1", "Columnas"],
        ["0", "Nada — primero, en el v3.1 escribe las cabeceras nuevas",
         "AF1 a AO1 (ver hoja CABECERAS NUEVAS)", "10 cabeceras"],
    ]
    col = 3
    for i, (dest, cabs) in enumerate(BLOQUES, 1):
        ini, fin = letra(col), letra(col + len(cabs) - 1)
        pasos.append([str(i),
                      f"{ini}2:{fin}{len(filas) + 1}   ({', '.join(cabs)})",
                      f"pegar en la celda {dest}{FILA_INICIAL}",
                      f"{len(cabs)} col."])
        col += len(cabs) + 1
    pasos += [
        [], ["Antes de pegar, comprueba que la columna B de la hoja DATOS coincide "
             "con la columna B del v3.1 en las filas 1989 a 2500."],
        ["Pega siempre con Pegado especial > Valores, para no arrastrar formato."],
        [f"Son {len(filas)} filas: de la {FILA_INICIAL} a la {FILA_INICIAL + len(filas) - 1} del v3.1."],
        ["Las celdas vacías corresponden a 44 cámaras que no aparecen en el proyecto; "
         "se dejan en blanco a propósito, en vez de #N/A."],
    ]
    for i, fila in enumerate(pasos, 1):
        for j, v in enumerate(fila, 1):
            c = ws.cell(i, j, v)
            if i == 1:
                c.font, c.fill, c.border = Font(bold=True), AMARILLO, FINO
            elif len(fila) == 1:
                c.font = Font(italic=True)
            c.alignment = Alignment(horizontal="left", vertical="center")
    for j, w in enumerate([8, 52, 34, 12], 1):
        ws.column_dimensions[letra(j)].width = w

    # ---- cabeceras nuevas para copiar tal cual
    ws = wb.create_sheet("CABECERAS NUEVAS")
    ws.cell(1, 1, "Copia esta fila y pégala en AF1 del v3.1").font = Font(bold=True, italic=True)
    for j, h in enumerate(BLOQUES[-1][1], 1):
        c = ws.cell(3, j, h)
        c.font, c.fill, c.border = Font(bold=True), AMARILLO, FINO
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[letra(j)].width = 18
    ws.cell(4, 1, f"↑ van en las columnas AF a AO ({letra(32)} … {letra(31 + len(BLOQUES[-1][1]))})")

    # ---- datos
    ws = wb.create_sheet("DATOS")
    ws.cell(1, 1, "Fila v3.1").fill = GRIS
    ws.cell(1, 2, "REFERENCIA (solo comprobación)").fill = GRIS
    for j in (1, 2):
        ws.cell(1, j).font, ws.cell(1, j).border = Font(bold=True), FINO
        ws.cell(1, j).alignment = Alignment(horizontal="center", wrap_text=True)
    col = 3
    mapa = []
    for dest, cabs in BLOQUES:
        for k, h in enumerate(cabs):
            c = ws.cell(1, col + k, h)
            c.font, c.fill, c.border = Font(bold=True), VERDE, FINO
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[letra(col + k)].width = 20
            mapa.append((col + k, h))
        ws.cell(2, col).comment = None
        col += len(cabs) + 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 26
    for i, fila in enumerate(filas, 2):
        ws.cell(i, 1, FILA_INICIAL + i - 2).alignment = Alignment(horizontal="center")
        ws.cell(i, 2, fila[cab["REFERENCIA"] - 1]).alignment = Alignment(horizontal="left")
        for j, h in mapa:
            c = ws.cell(i, j, fila[cab[h] - 1])
            c.alignment = Alignment(horizontal="center")
            c.border = FINO
    ws.freeze_panes = "C2"

    wb.save(salida)
    print(f"{len(filas)} filas | bloques: " +
          ", ".join(f"{d}{FILA_INICIAL} ({len(c)} col.)" for d, c in BLOQUES))


if __name__ == "__main__":
    main()
