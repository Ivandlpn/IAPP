#!/usr/bin/env python3
"""
SWITCH y BOCA de ALBALI a partir de las tablas de rotulación de cableado del
documento de parametrización del sistema de vídeo (ALBALI-AV-VCA-M700-SAL-008-A),
que vienen como imagen y hay que leer con OCR.

La etiqueta del switch sigue la regla documentada 6AA + tipo + emplazamiento +
SVC (switch) o RVC (router), así que la base se deriva del nombre de la cámara y
del OCR solo se toma el sufijo. El puerto no se puede derivar: sale del OCR.
"""
import json, re, sys, os, collections
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

INICIO, FIN = 1989, 2500
VERDE = PatternFill("solid", fgColor="FFD9EAD3")
GRIS = PatternFill("solid", fgColor="FFEFEFEF")
AZUL = PatternFill("solid", fgColor="FFDCE6F1")
FINO = Border(*[Side(style="thin")] * 4)
letra = lambda i: chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)
base = lambda ref: "6AA" + ref.split("_")[1] + ref.split("_")[2]


def arreglar_puerto(puerto, modelo):
    """Los MATCH-102 son modulares y su puerto va como «modulo,puerto». Cuando el
    OCR se come la coma queda «23» donde debería poner «2,3»."""
    if puerto and modelo and "MATCH" in modelo and "," not in puerto and len(puerto) == 2:
        return f"{puerto[0]},{puerto[1]}"
    return puerto


def main():
    datos, v31, salida = sys.argv[1], sys.argv[2], sys.argv[3]
    ocr = json.load(open(os.path.join(datos, "alb_switch_boca.json")))
    red = set(re.findall(r"6AA[A-Z0-9]+(?:SVC|RVC)", open(os.path.join(datos, "alb_RED.txt")).read()))

    # sufijo por emplazamiento: solo lecturas donde el switch leído coincide con
    # la base derivada del nombre de la cámara, que son las de fiar
    votos = collections.defaultdict(collections.Counter)
    for r, v in ocr.items():
        if v["switch"] and v["switch"][:-3] == base(r):
            votos[base(r)][v["switch"][-3:]] += 1
    suf_ocr = {s: c.most_common(1)[0][0] for s, c in votos.items()}
    suf_red = {}
    for e in red:
        suf_red.setdefault(e[:-3], e[-3:])

    ws31 = openpyxl.load_workbook(v31, data_only=True).active
    filas = []
    for r in range(INICIO, FIN + 1):
        ref = str(ws31.cell(r, 2).value).strip()
        actual_sw, actual_boca = ws31.cell(r, 24).value, ws31.cell(r, 27).value
        b = base(ref)
        if b in suf_ocr:
            sw, origen = b + suf_ocr[b], "leído en el PDF"
        elif b in suf_red:
            sw, origen = b + suf_red[b], "doc. de red"
        else:
            sw, origen = None, "sin evidencia"
        v = ocr.get(ref, {})
        puerto = arreglar_puerto(v.get("puerto"), v.get("modelo"))
        filas.append([
            r, ref,
            sw or actual_sw,                                   # X  SWITCH
            puerto or actual_boca,                             # AA BOCA
            origen,
            "referencia leída" if v.get("via") == "exacta" else
            ("referencia por parecido — revisar" if v.get("via") else "no leída"),
            v.get("n"), v.get("modelo"), v.get("pag"),
            b + "SVC" if origen == "sin evidencia" else None,
        ])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PEGAR EN EL v3.1"
    cab = ["Fila v3.1", "REFERENCIA (comprobación)",
           f"C → pegar en X{INICIO}\nSWITCH", f"D → pegar en AA{INICIO}\nBOCA",
           "solo consulta\nOrigen del switch", "solo consulta\nFiabilidad de la boca",
           "solo consulta\nVeces leída", "solo consulta\nModelo del switch",
           "solo consulta\nPág. PDF", "solo consulta\nSwitch por convención (sin confirmar)"]
    anchos = [9, 26, 20, 12, 18, 24, 12, 20, 10, 26]
    for j, h in enumerate(cab, 1):
        c = ws.cell(1, j, h)
        c.font = Font(bold=True)
        c.fill = GRIS if j <= 2 else (AZUL if j >= 5 else VERDE)
        c.border = FINO
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[letra(j)].width = anchos[j - 1]
    ws.row_dimensions[1].height = 46
    for i, fila in enumerate(filas, 2):
        for j, val in enumerate(fila, 1):
            c = ws.cell(i, j, val)
            c.border = FINO
            c.alignment = Alignment(horizontal="left" if j <= 2 else "center")
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{letra(len(cab))}{len(filas) + 1}"
    wb.save(salida)
    print(f"{len(filas)} filas | con SWITCH: {sum(1 for f in filas if f[2])} "
          f"| con BOCA: {sum(1 for f in filas if f[3])} "
          f"| bocas por parecido: {sum(1 for f in filas if 'parecido' in f[5])}")


if __name__ == "__main__":
    main()
