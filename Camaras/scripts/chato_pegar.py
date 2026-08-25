#!/usr/bin/env python3
"""
Cruza el bloque CHATO del listado v3.1 (filas 4468-4625) con el plan de
direccionamiento IP del sistema VCA y genera el fichero para pegar.
Todo va a columnas que ya existen.
"""
import json, os, sys
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

INICIO, FIN = 4468, 4625
VRM = "VRM Nuevo Chamartín"          # servidor de gestión de grabaciones del tramo
NVR_ANILLO = {1: "NVR-1 / NVR-2", 2: "NVR-3"}

VERDE = PatternFill("solid", fgColor="FFD9EAD3")
GRIS = PatternFill("solid", fgColor="FFEFEFEF")
AZUL = PatternFill("solid", fgColor="FFDCE6F1")
FINO = Border(*[Side(style="thin")] * 4)
letra = lambda i: chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)
es_na = lambda v: str(v).strip() == "#N/A"


def final(nuevo, actual):
    if nuevo not in (None, ""):
        return nuevo
    return None if (actual in (None, "") or es_na(actual)) else actual


def main():
    datos, v31, salida = sys.argv[1], sys.argv[2], sys.argv[3]
    recs = json.load(open(os.path.join(datos, "chato_ip.json")))
    por_ip = {r["ip"]: r for r in recs if r["servicio"] == "SERVICIO VIDEOVIGILANCIA"}

    ws31 = openpyxl.load_workbook(v31, data_only=True).active
    filas = []
    for r in range(INICIO, FIN + 1):
        v = [ws31.cell(r, c).value for c in range(1, 32)]
        p = por_ip.get(str(v[20]))
        nombre = v[2]                       # hoy el nombre del equipo vive en DESCRIPCIÓN
        filas.append([
            r, nombre,
            final(nombre, v[1]),                                  # B  REFERENCIA
            final(nombre, v[7]),                                  # H  NOMBRE DISPOSITIVO
            final("SEGURIDAD", v[18]),                            # S  FUNCIÓN CÁMARA
            final(p["puerto"] if p else None, v[26]),             # AA BOCA
            final(VRM, v[28]),                                    # AC SERVIDOR DE GRABACIÓN
            NVR_ANILLO.get(v[27]) if p else None,
            p["emplazamiento"] if p else None,
            p["equipo"] if p else None,
        ])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PEGAR EN EL v3.1"
    cab = ["Fila v3.1", "NOMBRE (comprobación)",
           f"C → pegar en B{INICIO}\nREFERENCIA",
           f"D → pegar en H{INICIO}\nNOMBRE DISPOSITIVO",
           f"E → pegar en S{INICIO}\nFUNCIÓN CÁMARA",
           f"F → pegar en AA{INICIO}\nBOCA",
           f"G → pegar en AC{INICIO}\nSERVIDOR DE GRABACIÓN",
           "solo consulta\nNVR del anillo", "solo consulta\nEmplazamiento (proyecto)",
           "solo consulta\nEquipo (proyecto)"]
    anchos = [9, 26, 26, 26, 16, 20, 24, 16, 28, 14]
    for j, h in enumerate(cab, 1):
        c = ws.cell(1, j, h)
        c.font = Font(bold=True)
        c.fill = GRIS if j <= 2 else (AZUL if j >= 8 else VERDE)
        c.border = FINO
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[letra(j)].width = anchos[j - 1]
    ws.row_dimensions[1].height = 46
    for i, fila in enumerate(filas, 2):
        for j, val in enumerate(fila, 1):
            c = ws.cell(i, j, val)
            c.border = FINO
            c.alignment = Alignment(horizontal="left" if j <= 2 else "center")
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{letra(len(cab))}{len(filas) + 1}"
    wb.save(salida)
    print(f"{len(filas)} filas | con BOCA: {sum(1 for f in filas if f[5])}")


if __name__ == "__main__":
    main()
