#!/usr/bin/env python3
"""
Vuelca la red de grabadores de TOVAL (documento LAVML-VIVS-HRD-RvidArqParamRedGrab_05)
a un libro listo para incorporar al listado general:

  - Grabadores TOVAL : los 107 grabadores con IP, máscara, gateway, nodo y VRM
  - Codec-Grabador   : a qué grabador va cada codificador (principal y fail-over)
  - Servidores VRM   : los tres servidores de gestión
  - PEGAR en cámaras : bloque alineado con las filas 2501-4467 del v3.1
"""
import json, sys, os, collections
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

FILA_INICIAL = 2501          # primera fila del bloque TOVAL en el v3.1
FILA_FINAL = 4467

AMARILLO = PatternFill("solid", fgColor="FFFFF200")
VERDE = PatternFill("solid", fgColor="FFD9EAD3")
GRIS = PatternFill("solid", fgColor="FFEFEFEF")
FINO = Border(*[Side(style="thin")] * 4)

sitio = lambda s: "_".join(str(s).split("_")[:3])
letra = lambda i: chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)

TRAMO = {"A6M": "Torrejón-Motilla [N1-N40]",
         "A6V": "Motilla-Valencia [N41-N70]",
         "A6A": "Motilla-Albacete [N71-N84]"}


def hoja(wb, titulo, cabeceras, filas, anchos, relleno=VERDE):
    ws = wb.create_sheet(titulo)
    for j, h in enumerate(cabeceras, 1):
        c = ws.cell(1, j, h)
        c.font, c.fill, c.border = Font(bold=True), relleno, FINO
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[letra(j)].width = anchos[j - 1]
    for i, fila in enumerate(filas, 2):
        for j, v in enumerate(fila, 1):
            c = ws.cell(i, j, v)
            c.border = FINO
            c.alignment = Alignment(horizontal="left" if j <= 2 else "center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{letra(len(cabeceras))}{len(filas) + 1}"
    return ws


def main():
    datos, v31, salida = sys.argv[1], sys.argv[2], sys.argv[3]
    G = json.load(open(os.path.join(datos, "toval_grabadores.json")))
    A = json.load(open(os.path.join(datos, "toval_asignacion.json")))
    sit2gr, gr_cod, gr_res = A["sit2gr"], A["gr_cod"], A["gr_res"]

    vrm_de = {g: d["vrm"] for g, d in G.items()}
    vrm_sitio = {}
    for s, gs in sit2gr.items():
        v = {vrm_de[g] for g in gs if g in vrm_de}
        if len(v) == 1:
            vrm_sitio[s] = v.pop()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CÓMO AÑADIRLO"
    texto = [
        ("Red de grabadores de TOVAL — cómo incorporarla al listado general", True),
        ("", False),
        ("1. Copia la hoja «Grabadores TOVAL» entera al libro del listado general, "
         "como hoja nueva.", False),
        ("   Es información de equipo de grabación, no de cámara: no cabe como columnas "
         "en la hoja de cámaras.", False),
        ("", False),
        ("2. Haz lo mismo con «Codec-Grabador» y «Servidores VRM».", False),
        ("", False),
        ("3. En la hoja «PEGAR en cámaras» tienes un bloque alineado con las filas "
         f"{FILA_INICIAL} a {FILA_FINAL} del v3.1 (las de TOVAL).", False),
        ("   Copia C2:C1968 y pégalo en AC" + str(FILA_INICIAL) +
         "  (columna SERVIDOR DE GRABACIÓN, hoy vacía en TOVAL).", False),
        ("   Copia D2:D1968 y pégalo en la primera columna libre, con cabecera «VRM».", False),
        ("   Pega siempre con Pegado especial > Valores.", False),
        ("", False),
        ("LIMITACIÓN IMPORTANTE", True),
        ("El documento asigna grabador por CODIFICADOR, no por cámara. La columna "
         "CODIFICADOR/PUERTO de TOVAL está vacía en el listado, así que el enlace "
         "cámara → grabador solo se puede cerrar del todo cuando se rellene.", False),
        ("Por eso la columna SERVIDOR DE GRABACIÓN solo viene rellena donde el "
         "emplazamiento tiene un único grabador. En el resto se deja en blanco y los "
         "candidatos quedan en la columna E, a título informativo.", False),
        ("", False),
        ("Cuando CODIFICADOR/PUERTO esté relleno, la hoja «Codec-Grabador» cierra el "
         "enlace con un BUSCARV directo.", False),
    ]
    for i, (t, negrita) in enumerate(texto, 1):
        c = ws.cell(i, 1, t)
        c.font = Font(bold=negrita, size=12 if negrita else 11)
        c.alignment = Alignment(horizontal="left")
    ws.column_dimensions["A"].width = 110

    hoja(wb, "Grabadores TOVAL",
         ["GRABADOR", "EMPLAZAMIENTO", "NODO", "TRAMO", "DIRECCIÓN IP",
          "MÁSCARA DE RED", "PUERTA DE ENLACE", "VRM", "Nº CODECS ASIGNADOS",
          "Nº CODECS FAIL-OVER", "Pág. PDF"],
         sorted([[g, d["desc"], d["nodo"], TRAMO.get(g[:3], ""), d["ip"], d["mask"],
                  d["gw"], int(d["vrm"]) if d["vrm"] else None,
                  len(gr_cod.get(g, [])), len(gr_res.get(g, [])), d["pag"]]
                 for g, d in G.items()]),
         [22, 34, 8, 26, 16, 18, 18, 8, 14, 14, 10])

    codecs = []
    for g, cs in sorted(gr_cod.items()):
        for c in cs:
            codecs.append([c, g, G.get(g, {}).get("desc"),
                           int(G[g]["vrm"]) if g in G and G[g]["vrm"] else None, "principal"])
    for g, cs in sorted(gr_res.items()):
        for c in cs:
            codecs.append([c, g, G.get(g, {}).get("desc"),
                           int(G[g]["vrm"]) if g in G and G[g]["vrm"] else None, "fail-over"])
    hoja(wb, "Codec-Grabador",
         ["CODIFICADOR", "GRABADOR", "EMPLAZAMIENTO DEL GRABADOR", "VRM", "PAPEL"],
         sorted(codecs), [22, 22, 34, 8, 14])

    hoja(wb, "Servidores VRM",
         ["VRM", "QUÉ GRABA", "Nº DVA i-SCSI"],
         [[1, "Tramo Motilla-Valencia [N41-N70], Tramo Motilla-Albacete [N71-N84], "
              "SSEE Torrent", 37],
          [2, "Tramo Torrejón-Motilla [N1-N40]", 36],
          [3, "Estaciones de Cuenca, Requena, Albacete y Valencia; Bases de Mantenimiento "
              "de Villarrubia, Gabaldón, Requena y La Gineta; CRC Albacete", 33]],
         [8, 80, 14])

    # bloque para pegar
    wsx = openpyxl.load_workbook(v31, data_only=True).active
    filas = []
    for r in range(FILA_INICIAL, FILA_FINAL + 1):
        ref = wsx.cell(r, 2).value
        s = sitio(ref)
        gs = sit2gr.get(s, [])
        filas.append([r, ref,
                      gs[0] if len(gs) == 1 else None,
                      int(vrm_sitio[s]) if s in vrm_sitio else None,
                      " / ".join(gs) if len(gs) > 1 else None])
    ws = hoja(wb, "PEGAR en cámaras",
              ["Fila v3.1", "REFERENCIA (solo comprobación)",
               f"SERVIDOR DE GRABACIÓN → pegar en AC{FILA_INICIAL}",
               "VRM → columna nueva", "Grabadores candidatos (informativo)"],
              filas, [10, 28, 30, 16, 40], GRIS)
    for j in (3, 4):
        ws.cell(1, j).fill = VERDE
    ws.freeze_panes = "C2"

    wb.save(salida)
    print(f"grabadores: {len(G)} | codecs: {len(codecs)} | filas TOVAL: {len(filas)} | "
          f"con grabador: {sum(1 for f in filas if f[2])} | con VRM: {sum(1 for f in filas if f[3])}")


if __name__ == "__main__":
    main()
