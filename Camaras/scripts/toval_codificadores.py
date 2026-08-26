#!/usr/bin/env python3
"""
Cierra la cadena cámara → codificador → grabador de TOVAL cruzando:

  - Red de Codificadores (RvidArqParamRedCodif_06): qué cámara va a qué
    codificador y en qué entrada.
  - Red de Grabadores (RvidArqParamRedGrab_05): qué grabador graba cada
    codificador, como principal y como fail-over.

El cruce con el listado se hace por nombre de equipo y, cuando el listado usa
otra nomenclatura, por ubicación más sufijo de cámara.
"""
import json, os, re, sys, unicodedata, collections
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

INICIO, FIN = 2501, 4467
VERDE = PatternFill("solid", fgColor="FFD9EAD3")
GRIS = PatternFill("solid", fgColor="FFEFEFEF")
AZUL = PatternFill("solid", fgColor="FFDCE6F1")
FINO = Border(*[Side(style="thin")] * 4)
letra = lambda i: chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)
K = lambda s: re.sub(r"[^A-Z0-9]", "", str(s).upper())


def N(s):
    s = str(s or "").strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return re.sub(r"[^A-Z0-9]", "", s)


def main():
    datos, v31, salida = sys.argv[1], sys.argv[2], sys.argv[3]
    asig = json.load(open(os.path.join(datos, "tov_codif_asig.json")))
    cad = json.load(open(os.path.join(datos, "tov_cod2gr.json")))
    ips = json.load(open(os.path.join(datos, "tov_codec_ip.json")))
    cod2gr, cod2res = cad["cod2gr"], cad["cod2res"]

    # codecs que el documento de grabadores lista bajo más de un grabador
    # principal: la lectura no es concluyente y conviene avisar
    ambiguos = set(cad.get("ambiguos", []))

    cams = [a for a in asig if a["cam"].startswith("A6")]
    por_nombre, por_sitio = {}, {}
    for a in cams:
        por_nombre.setdefault(K(a["cam"]), a)
        por_sitio.setdefault((N(a["desc"]), a["cam"].split("_")[-1]), a)

    ws31 = openpyxl.load_workbook(v31, data_only=True).active
    filas = []
    for r in range(INICIO, FIN + 1):
        ref = str(ws31.cell(r, 2).value).strip()
        ubic = ws31.cell(r, 13).value
        a = por_nombre.get(K(ref))
        via = "nombre" if a else None
        if not a:
            a = por_sitio.get((N(ubic), ref.split("_")[-1]))
            via = "ubicación + sufijo" if a else "no encontrada"
        codec = a["codec"] if a else None
        grab = cod2gr.get(codec) if codec else None
        res = cod2res.get(codec) if codec else None
        filas.append([
            r, ref,
            "ANALÓGICA",                                   # D  TECNOLOGÍA
            codec,                                         # E  CODIFICADOR/PUERTO
            a["ent"] if a else None,                       # F  PUERTO
            ref,                                           # H  NOMBRE DISPOSITIVO
            "SEGURIDAD",                                   # S  FUNCIÓN CÁMARA
            grab,                                          # AC SERVIDOR DE GRABACIÓN
            " / ".join(res) if res else None,
            (ips.get(codec) or {}).get("ip"),
            a["cam"] if a and a["cam"] != ref else None,
            a["desc"] if a else None,
            via,
            "revisar: el codec figura bajo varios grabadores" if codec in ambiguos else None,
        ])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CÓMO PEGARLO"
    lineas = [
        ("TOVAL: cámara → codificador → grabador", True),
        ("", False),
        ("Filas 2501 a 4467 del v3.1. Cinco pegados, todos sobre columnas que ya existen:", False),
        ("", False),
        ("   C2:C1968   →  D2501    TECNOLOGÍA = ANALÓGICA", False),
        ("   D2:E1968   →  E2501    CODIFICADOR/PUERTO + PUERTO   (dos columnas de una vez)", False),
        ("   F2:F1968   →  H2501    NOMBRE DISPOSITIVO", False),
        ("   G2:G1968   →  S2501    FUNCIÓN CÁMARA = SEGURIDAD", False),
        ("   H2:H1968   →  AC2501   SERVIDOR DE GRABACIÓN", False),
        ("", False),
        ("Pegado especial > Valores. Antes, comprueba que la columna B de la hoja DATOS "
         "coincide con la columna B del v3.1 en esas filas.", False),
        ("", False),
        ("Las columnas azules no se pegan: son el grabador de fail-over, la IP del "
         "codificador, el nombre que usa el proyecto cuando difiere del listado, el "
         "emplazamiento y cómo se ha cruzado cada fila.", False),
    ]
    for i, (t, b) in enumerate(lineas, 1):
        c = ws.cell(i, 1, t)
        c.font = Font(bold=b, size=12 if b else 11)
        c.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 108

    ws = wb.create_sheet("DATOS")
    cab = ["Fila v3.1", "REFERENCIA (comprobación)",
           f"C → D{INICIO}\nTECNOLOGÍA", f"D → E{INICIO}\nCODIFICADOR/PUERTO",
           f"E → F{INICIO}\nPUERTO", f"F → H{INICIO}\nNOMBRE DISPOSITIVO",
           f"G → S{INICIO}\nFUNCIÓN CÁMARA", f"H → AC{INICIO}\nSERVIDOR DE GRABACIÓN",
           "solo consulta\nGrabador fail-over", "solo consulta\nIP del codificador",
           "solo consulta\nNombre en el proyecto", "solo consulta\nEmplazamiento",
           "solo consulta\nCruzado por", "solo consulta\nAviso"]
    anchos = [9, 24, 14, 24, 10, 24, 16, 24, 24, 16, 24, 30, 16, 34]
    for j, h in enumerate(cab, 1):
        c = ws.cell(1, j, h)
        c.font = Font(bold=True)
        c.fill = GRIS if j <= 2 else (VERDE if j <= 8 else AZUL)
        c.border = FINO
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[letra(j)].width = anchos[j - 1]
    ws.row_dimensions[1].height = 46
    for i, fila in enumerate(filas, 2):
        for j, val in enumerate(fila, 1):
            c = ws.cell(i, j, val)
            c.border = FINO
            c.alignment = Alignment(horizontal="left" if j <= 2 else "center")
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{letra(len(cab))}{len(filas) + 1}"

    wb.save(salida)
    print(f"{len(filas)} filas | codificador: {sum(1 for f in filas if f[3])} | "
          f"grabador: {sum(1 for f in filas if f[7])} | "
          f"fail-over: {sum(1 for f in filas if f[8])} | "
          f"avisos: {sum(1 for f in filas if f[13])}")


if __name__ == "__main__":
    main()
